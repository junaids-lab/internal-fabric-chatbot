import argparse
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract RDCCI Internal chatbot KB docs from workbook metadata.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    docs = []
    docs.extend(extract_table_column_docs(required_sheet(wb, "Final-Columns with Lookup"), workbook_path.name))
    docs.extend(extract_question_docs(required_sheet(wb, "Questions"), workbook_path.name))

    measures_sheet = find_sheet(wb, "Measures Semantic Model", "Measures created in semantic models")
    if measures_sheet is not None:
        docs.extend(extract_measure_docs(measures_sheet, workbook_path.name))

    general_mapping_sheet = find_sheet(wb, "General Mapping", "general mapping for chatbot")
    if general_mapping_sheet is not None:
        docs.extend(extract_general_mapping_docs(general_mapping_sheet, workbook_path.name))

    semantic_model_sheet = find_sheet(wb, "Semantic Model")
    if semantic_model_sheet is not None:
        docs.extend(extract_semantic_model_docs(semantic_model_sheet, workbook_path.name))

    table_names_sheet = find_sheet(wb, "Table Names", "table names")
    if table_names_sheet is not None:
        docs.extend(extract_table_definition_docs(table_names_sheet, workbook_path.name))

    final_data_tables_sheet = find_sheet(wb, "Final Data Tables", "final data tables")
    if final_data_tables_sheet is not None:
        docs.extend(extract_final_data_table_docs(final_data_tables_sheet, workbook_path.name))

    tables_attributes_sheet = find_sheet(wb, "Tables & Attributes", "tables and attributes")
    if tables_attributes_sheet is not None:
        docs.extend(extract_tables_attributes_docs(tables_attributes_sheet, workbook_path.name))

    with out_path.open("w", encoding="utf-8") as handle:
        for doc in docs:
            handle.write(json.dumps(doc, ensure_ascii=False) + "\n")

    print(f"Wrote {len(docs)} documents to {out_path}")


def extract_table_column_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []
    context: dict[str, Any] = {}

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        for key in [
            "dataflow_gen2_name",
            "status",
            "samentic_model",
            "source_table_name",
            "type",
            "table_type",
            "data_type",
        ]:
            if record.get(key):
                context[key] = record[key]

        column = record.get("source_column")
        if not column or not context.get("source_table_name"):
            continue

        semantic_model = normalize_semantic_model(str(context.get("samentic_model") or "").strip())
        table_name = str(context.get("source_table_name") or "").strip()
        column_name = str(column).strip()
        key_type = str(record.get("key") or "").strip()
        title = f"{semantic_model}.{table_name}.{column_name}"
        content = "\n".join(
            [
                f"Semantic model: {semantic_model}",
                f"Dataflow: {context.get('dataflow_gen2_name', '')}",
                f"Domain: {context.get('type', '')}",
                f"Table: {table_name}",
                f"Column: {column_name}",
                f"Key: {key_type}",
                f"Table type: {context.get('table_type', '')}",
                f"Data type: {context.get('data_type', '')}",
            ]
        )
        docs.append(
            {
                "id": stable_id("column", source_name, row_number, title),
                "doc_type": "table_column",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": table_name,
                "column_name": column_name,
                "language": "mixed",
            }
        )

    return docs


def extract_question_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []
    section = ""

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        if record.get("ai_query") and not record.get("samentic_layer"):
            section = str(record["ai_query"])
            continue

        ai_query = clean_text(record.get("ai_query"))
        semantic_model = normalize_semantic_model(clean_text(record.get("samentic_layer")))
        if not ai_query or not semantic_model:
            continue

        sql_statement = clean_text(record.get("sql_statement"))
        notes = clean_text(record.get("notes"))
        metadata = clean_text(record.get("metadata"))
        table_names = clean_text(record.get("table_names"))
        title = f"Sample question {record.get('sl_no') or row_number}: {ai_query[:80]}"
        content = "\n".join(
            part
            for part in [
                f"Section: {section}",
                f"Question: {ai_query}",
                f"Semantic model: {semantic_model}",
                f"Tables: {table_names}",
                f"Business metadata and synonyms: {metadata}",
                f"Notes: {notes}",
                f"Reference SQL from workbook: {sql_statement}",
            ]
            if part and not part.endswith(": ")
        )
        docs.append(
            {
                "id": stable_id("question", source_name, row_number, ai_query),
                "doc_type": "sample_question",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": table_names,
                "column_name": "",
                "language": "mixed",
            }
        )

    return docs


def extract_measure_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        semantic_model = normalize_semantic_model(clean_text(record.get("semantic_model")))
        measure_name = clean_text(record.get("measure_name"))
        if not semantic_model or not measure_name:
            continue

        date_column = clean_text(record.get("date_column"))
        dax_formula = clean_text(record.get("dax_formula"))
        title = f"{semantic_model} measure: {measure_name}"
        content = "\n".join(
            [
                f"Semantic model: {semantic_model}",
                f"Measure name: {measure_name}",
                f"Default date column: {date_column}",
                f"DAX formula: {dax_formula}",
            ]
        )
        docs.append(
            {
                "id": stable_id("measure", source_name, row_number, semantic_model, measure_name),
                "doc_type": "measure",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": "",
                "column_name": date_column,
                "language": "mixed",
            }
        )

    return docs


def extract_general_mapping_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        arabic_term = clean_text(record.get("arabic_user_term"))
        english_term = clean_text(record.get("english_term"))
        semantic_model = normalize_semantic_model(clean_text(record.get("semantic_model")))
        measure_name = clean_text(record.get("measure_name"))
        if not semantic_model or not measure_name or not (arabic_term or english_term):
            continue

        default_date_column = clean_text(record.get("default_date_column"))
        allowed_dimensions = clean_text(record.get("allowed_dimensions"))
        requires_clarification = clean_text(record.get("requires_clarification"))
        title = f"Business mapping: {arabic_term or english_term}"
        content = "\n".join(
            [
                f"Arabic user term: {arabic_term}",
                f"English term: {english_term}",
                f"Semantic model: {semantic_model}",
                f"Measure name: {measure_name}",
                f"Default date column: {default_date_column}",
                f"Allowed dimensions: {allowed_dimensions}",
                f"Requires clarification: {requires_clarification}",
            ]
        )
        docs.append(
            {
                "id": stable_id("mapping", source_name, row_number, arabic_term, english_term, measure_name),
                "doc_type": "business_mapping",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": allowed_dimensions,
                "column_name": default_date_column,
                "language": "mixed",
            }
        )

    return docs


def extract_semantic_model_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []
    context: dict[str, str] = {}

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        if record.get("semantic_model"):
            context["semantic_model"] = normalize_semantic_model(clean_text(record.get("semantic_model")))
        table_name = clean_text(record.get("tables"))
        semantic_model = context.get("semantic_model", "")
        if not semantic_model or not table_name:
            continue

        filter_note = clean_text(record.get("filter_based_on_the_column_a"))
        status = clean_text(record.get("status"))
        title = f"{semantic_model} table mapping: {table_name}"
        content = "\n".join(
            [
                f"Semantic model: {semantic_model}",
                f"Table: {table_name}",
                f"Filter note: {filter_note}",
                f"Status: {status}",
            ]
        )
        docs.append(
            {
                "id": stable_id("semantic-model-table", source_name, row_number, semantic_model, table_name),
                "doc_type": "semantic_model_table",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": table_name,
                "column_name": "",
                "language": "mixed",
            }
        )

    return docs


def extract_table_definition_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        semantic_model = normalize_semantic_model(first_value(record, "semantic_model", "samentic_model", "model", "dataset"))
        table_name = first_value(record, "table_name", "tables", "table", "source_table_name")
        definition = first_value(record, "definition", "description", "table_definition", "business_definition", "metadata", "notes")
        if not table_name and not definition:
            continue

        title = f"Table definition: {table_name or row_number}"
        content = generic_record_content(record, preferred_keys=["semantic_model", "table_name", "definition", "description", "metadata", "notes"])
        docs.append(
            {
                "id": stable_id("table-definition", source_name, row_number, semantic_model, table_name, definition),
                "doc_type": "table_definition",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": table_name,
                "column_name": "",
                "language": "mixed",
            }
        )

    return docs


def extract_final_data_table_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []
    context: dict[str, str] = {}

    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        semantic_model = normalize_semantic_model(first_value(record, "semantic_model", "samentic_model", "model", "dataset"))
        table_name = first_value(record, "table_name", "tables", "table", "source_table_name")
        column_name = first_value(record, "column_name", "source_column", "column", "field")

        if semantic_model:
            context["semantic_model"] = semantic_model
        if table_name:
            context["table_name"] = table_name

        semantic_model = semantic_model or context.get("semantic_model", "")
        table_name = table_name or context.get("table_name", "")
        if not table_name and not column_name:
            continue

        title = f"Final data table: {semantic_model}.{table_name}" if semantic_model else f"Final data table: {table_name}"
        if column_name:
            title = f"{title}.{column_name}"

        docs.append(
            {
                "id": stable_id("final-data-table", source_name, row_number, semantic_model, table_name, column_name),
                "doc_type": "final_data_table",
                "title": title,
                "content": generic_record_content(record),
                "source": source_name,
                "semantic_model": semantic_model,
                "table_name": table_name,
                "column_name": column_name,
                "language": "mixed",
            }
        )

    return docs


def extract_tables_attributes_docs(ws: Any, source_name: str) -> list[dict[str, Any]]:
    header, rows, start_row = rows_with_header(ws)
    if not rows:
        return []

    docs: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=start_row):
        record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
        table_name = first_value(record, "source_table_name", "table_name", "table")
        table_type = first_value(record, "table_type")
        data_type = first_value(record, "data_type")
        comments = first_value(record, "comments", "notes")
        attribute_values = [clean_text(value) for key, value in record.items() if key.startswith("column") and clean_text(value)]
        if not table_name and not attribute_values:
            continue

        title = f"Table attributes: {table_name or row_number}"
        content = "\n".join(
            part
            for part in [
                f"Table: {table_name}",
                f"Table type: {table_type}",
                f"Data type: {data_type}",
                f"Attributes: {', '.join(attribute_values)}",
                f"Comments: {comments}",
            ]
            if part and not part.endswith(": ")
        )
        docs.append(
            {
                "id": stable_id("table-attributes", source_name, row_number, table_name, ",".join(attribute_values)),
                "doc_type": "table_attributes",
                "title": title,
                "content": content,
                "source": source_name,
                "semantic_model": "",
                "table_name": table_name,
                "column_name": ", ".join(attribute_values),
                "language": "mixed",
            }
        )

    return docs


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("-", "_").replace(" ", "_")
    text = text.replace(".", "").replace("/", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


def rows_with_header(ws: Any) -> tuple[list[str], list[tuple[Any, ...]], int]:
    raw_rows = list(ws.iter_rows(values_only=True))
    for index, row in enumerate(raw_rows):
        if any(clean_text(value) for value in row):
            header = [normalize_header(value) for value in row]
            return header, raw_rows[index + 1 :], index + 2
    return [], [], 1


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def first_value(record: dict[str, Any], *keys: str) -> str:
    normalized_keys = [normalize_header(key) for key in keys]
    for key in normalized_keys:
        value = clean_text(record.get(key))
        if value:
            return value
    return ""


def generic_record_content(record: dict[str, Any], preferred_keys: list[str] | None = None) -> str:
    ordered_keys = []
    if preferred_keys:
        ordered_keys.extend(normalize_header(key) for key in preferred_keys)
    ordered_keys.extend(key for key in record if key not in ordered_keys)

    lines = []
    for key in ordered_keys:
        value = clean_text(record.get(key))
        if value:
            lines.append(f"{key.replace('_', ' ').title()}: {value}")
    return "\n".join(lines)


def normalize_semantic_model(value: str) -> str:
    normalized_values = []
    for part in value.replace(";", "\n").replace(",", "\n").splitlines():
        normalized = part.strip().replace("dddm_sm_manualtamp", "dddm_sm_manualstamp")
        if normalized and normalized not in normalized_values:
            normalized_values.append(normalized)
    return ", ".join(normalized_values)


def find_sheet(wb: Any, *names: str) -> Any | None:
    sheet_lookup = {normalize_header(name): name for name in wb.sheetnames}
    for name in names:
        sheet_name = sheet_lookup.get(normalize_header(name))
        if sheet_name:
            return wb[sheet_name]
    return None


def required_sheet(wb: Any, name: str) -> Any:
    sheet = find_sheet(wb, name)
    if sheet is None:
        raise KeyError(f"Workbook is missing required sheet: {name}")
    return sheet


def stable_id(*parts: Any) -> str:
    raw = "|".join(str(part) for part in parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]


if __name__ == "__main__":
    main()
