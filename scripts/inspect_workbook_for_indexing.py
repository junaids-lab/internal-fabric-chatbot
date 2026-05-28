import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl


INDEX_SHEET_HINTS = {
    "semantic_model",
    "final_columns_with_lookup",
    "table_names",
    "tables_&_attributes",
    "tables_and_attributes",
    "final_data_tables",
    "questions",
    "measures_created_in_semantic_models",
    "measures_semantic_model",
    "general_mapping_for_chatbot",
    "general_mapping",
}

TRANSACTIONAL_HINTS = {
    "amount",
    "balance",
    "customer",
    "member_name",
    "mobile",
    "email",
    "phone",
    "address",
    "national_id",
    "iqama",
    "passport",
    "transaction_id",
    "invoice",
    "voucher_no",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect workbook sheets before Azure AI Search indexing.")
    parser.add_argument("--workbook", required=True, help="Absolute path to the Excel workbook.")
    parser.add_argument("--out", default="docs/workbook_indexing_review.md", help="Markdown report output path.")
    parser.add_argument("--json-out", default="", help="Optional JSON report output path.")
    parser.add_argument("--sample-rows", type=int, default=3, help="Number of non-empty sample rows per sheet.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)

    report = {
        "workbook": str(workbook_path),
        "sheets": [inspect_sheet(wb[sheet_name], args.sample_rows) for sheet_name in wb.sheetnames],
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(render_markdown(report), encoding="utf-8")

    if args.json_out:
        json_out_path = Path(args.json_out)
        json_out_path.parent.mkdir(parents=True, exist_ok=True)
        json_out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Workbook sheets reviewed: {len(report['sheets'])}")
    print(f"Wrote report: {out_path}")


def inspect_sheet(ws: Any, sample_rows: int) -> dict[str, Any]:
    all_rows = list(ws.iter_rows(values_only=True))
    header_index = 0
    header_row = []
    for index, row in enumerate(all_rows):
        if any(clean_text(value) for value in row):
            header_index = index
            header_row = row
            break
    headers = [clean_text(value) for value in header_row]
    normalized_headers = [normalize_header(value) for value in headers]

    non_empty_rows = 0
    samples = []
    for row in all_rows[header_index + 1 :]:
        values = [clean_text(value) for value in row]
        if not any(values):
            continue
        non_empty_rows += 1
        if len(samples) < sample_rows:
            samples.append(sample_record(headers, values))

    normalized_sheet_name = normalize_header(ws.title)
    include_recommendation, reason = classify_sheet(normalized_sheet_name, normalized_headers)

    return {
        "sheet": ws.title,
        "normalized_sheet": normalized_sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "non_empty_data_rows": non_empty_rows,
        "headers": headers,
        "normalized_headers": normalized_headers,
        "recommendation": include_recommendation,
        "reason": reason,
        "sample_rows": samples,
    }


def classify_sheet(normalized_sheet_name: str, normalized_headers: list[str]) -> tuple[str, str]:
    header_set = {header for header in normalized_headers if header}

    if normalized_sheet_name in INDEX_SHEET_HINTS:
        return "include", "Known metadata/routing workbook sheet."

    if header_set & TRANSACTIONAL_HINTS:
        return "review_or_skip", "Headers suggest possible transactional or sensitive row-level data."

    if {"semantic_model", "table_name"} & header_set or {"measure_name", "dax_formula"} & header_set:
        return "include", "Headers suggest semantic model, table, or measure metadata."

    if {"arabic_user_term", "english_term"} & header_set:
        return "include", "Headers suggest chatbot terminology mapping."

    return "review", "Unknown sheet shape; review samples before indexing."


def sample_record(headers: list[str], values: list[str]) -> dict[str, str]:
    record = {}
    for index, value in enumerate(values):
        if not value:
            continue
        header = headers[index] if index < len(headers) and headers[index] else f"Column {index + 1}"
        record[header] = value[:300]
    return record


def render_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Workbook Indexing Review",
        "",
        f"Workbook: `{report['workbook']}`",
        "",
        "| Sheet | Rows | Columns | Recommendation | Reason |",
        "| --- | ---: | ---: | --- | --- |",
    ]

    for sheet in report["sheets"]:
        lines.append(
            "| "
            f"{sheet['sheet']} | "
            f"{sheet['non_empty_data_rows']} | "
            f"{sheet['max_column']} | "
            f"{sheet['recommendation']} | "
            f"{sheet['reason']} |"
        )

    for sheet in report["sheets"]:
        lines.extend(
            [
                "",
                f"## {sheet['sheet']}",
                "",
                f"Recommendation: **{sheet['recommendation']}**",
                "",
                f"Reason: {sheet['reason']}",
                "",
                "Headers:",
                "",
                ", ".join(header for header in sheet["headers"] if header) or "No headers detected.",
                "",
                "Sample rows:",
                "",
                "```json",
                json.dumps(sheet["sample_rows"], ensure_ascii=False, indent=2),
                "```",
            ]
        )

    lines.append("")
    return "\n".join(lines)


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_header(value: Any) -> str:
    text = clean_text(value).lower()
    text = text.replace("\n", " ").replace("-", "_").replace(" ", "_")
    text = text.replace(".", "").replace("/", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


if __name__ == "__main__":
    main()